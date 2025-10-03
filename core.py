import random
import os
import math

from openpyxl import load_workbook, Workbook

year = 2025

def month_creator(month, year):
    """Make exel file, make peoples - samples with random birth years and month."""
    if os.path.exists('years_ago_'+str(j)+'.xlsx'):
        pass
    else:
        wb = Workbook()
        wb.save('years_ago_'+str(j)+'.xlsx')
        wb.close()

    wb = load_workbook('years_ago_'+str(j)+'.xlsx')
    ws = wb.active

    for i in range(1, 10000):
        birth_year = random.randint(1900, 2010)
        birth_month = random.randint(1, 12)

        ws[f'A{i}'] = birth_year 
        ws[f'B{i}'] = birth_month

        if birth_month < month:
            ws[f'C{i}'] = year - birth_year
        else:
            ws[f'C{i}'] = year - birth_year - 1

    wb.save('years_ago_'+str(j)+'.xlsx')

def correlation_calculator(file):
    """Calculation of corelation coefficient for birth month and birth years of first 10000 samples"""
    year_sum = 0
    age_sum = 0

    wb = load_workbook(file)
    ws = wb.active

    for i in range(1, 10000):
        year_sum += ws[f'A{i}'].value
        age_sum += ws[f'C{i}'].value

    year_mean = year_sum/9999
    age_mean = age_sum/9999

    year_sko =  0
    age_sko = 0

    for i in range(1, 10000):
        year_sko += (ws[f'A{i}'].value-year_mean)**2
        age_sko += (ws[f'C{i}'].value-age_mean)**2

    year_sko = math.sqrt(year_sko/(9999-1))
    age_sko = math.sqrt(age_sko/(9999-1))

    sum_pq = 0

    for i in range(1, 10000):
        sum_pq += (ws[f'A{i}'].value*ws[f'C{i}'].value)

    correlation = (sum_pq-9999*year_mean*age_mean)/((9999-1)*year_sko*age_sko)

    print(f'Корелляция для {file} = {correlation}')

def get_sign(filename, n_samples):
    """Method get name of file and number of n_samples and returns list of signs"""
    wb = load_workbook(filename)
    ws = wb.active

    sign_list = []

    for i in range(1, n_samples+1):
        sign_list.append(int(ws[f'C{i}'].value))
    
    return sign_list

def sort_list(unsort_list):
    """The method takes an unsorted list at input and returns a sorted list"""
    sort_list = [unsort_list[0]]

    for i in range(1, len(unsort_list)):
        for j in range(len(sort_list)):
            if unsort_list[i] > sort_list[j]:
                sort_list.insert(j, unsort_list[i])
                break

    return sort_list

def get_median(sort_list):
    """Method input is sorted list of values and return is median"""
    if len(sort_list)%2 == 1:
        median = sort_list[int(len(sort_list)//2)]
    else:
        median = (sort_list[int(len(sort_list)//2-1)] + sort_list[int(len(sort_list)//2)])/2

    return median

def mean(list_):
    """Calculate mean value and return float"""
    sum = 0
    for i in list_:
        sum += i
    return sum/len(list_)

def sko(list_):
    """Calculate RSD and return float"""
    sum_for_sko = 0
    list_mean = mean(list_)
    for i in list_:
        sum_for_sko += (i-list_mean)**2
    return math.sqrt(sum_for_sko/(len(list_)-1))

def signs_normalize(filename, data_sheet, normilized_sheet):
    #Normalizing data p-pmean/sko and return list of list of signs of ever sample 
    ws = load_workbook(filename)
    wb = ws[data_sheet]
    column = 1
    signs = []
    while wb.cell(row=1, column=column).value != None:
        row = 1
        sign = []
        while wb.cell(row=row, column=column).value != None:
            sign.append(wb.cell(row=row, column=column).value)
            row+=1
        
        column += 1
        signs.append(sign)

    wb_normalized = ws[normilized_sheet]

    for i in range(len(signs)):
        sign_mean = mean(signs[i])
        sign_sko = sko(signs[i])
        for j in range(len(signs[i])):
            wb_normalized.cell(column=i+1, row=j+1, value=(signs[i][j]-sign_mean)/sign_sko)

    ws.save(filename)
    return signs

def mae_calc(values_list, predict_values_list):
    """Function calculate mae and return float"""
    if len(values_list) != len(predict_values_list):
        raise Exception("Number of values and prediction values not equal")

    mae = 0
    for i in range(len(values_list)):
        mae += (abs(values_list[i]-predict_values_list[i]))

    return mae

def mape_calc(values_list, predict_values_list):
    """Function calculate mape and return float"""
    if len(values_list) != len(predict_values_list):
        raise Exception("Number of values and prediction values not equal")

    mape = 0
    for i in range(len(values_list)):
        mape += ((abs(values_list[i]-predict_values_list[i]))/abs(values_list[i]))
    mape = mape*100
    mape = mape/len(values_list)
    return mape

#for j in range(1, 13):
#    month_creator(j, year)
#    correlation_calculator(f'years_ago_{j}.xlsx')