from openpyxl import load_workbook

from vector import Vector

def load_dataset(filename, sheet, start_column = 1, start_row = 1):
    """Load dataset from exel and return two lists"""
    ws = load_workbook(filename)
    wb = ws[sheet]

    coords = []
    row = start_row
    while wb.cell(row=row, column=start_column).value is not None:
        sample_coords = []
        column = start_column
        while wb.cell(row=row, column=column).value is not None:
            sample_coords.append(wb.cell(row=row, column=column).value)
            column += 1
        coords.append(sample_coords)
        row += 1

    vectors = []
    for i in coords:
        vectors.append(Vector(i))
    
    return coords, vectors