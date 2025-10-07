from openpyxl import load_workbook

from load import load_dataset
from vector import Vector
from forel import get_cluster_vectors
from core import mape_calc
from matrix import Matrix

filename = 'Test.xlsx'
sheet = 'Data'

wb = load_workbook(filename)
ws = wb[sheet]

X = []
i = 1
col = 2
while ws.cell(row = 1, column = col).value != None:
    X_n_vals = []
    while ws.cell(row = i, column = 1).value != None:
        X_n_vals.append(ws.cell(row = i, column = 1).value)
        i += 1

    X.append(X_n_vals)
    col += 1

Y = [[]]

i = 1
while ws.cell(row = i, column = 2).value != None:
    Y[0].append(ws.cell(row = i, column = 2).value)
    i += 1

X= Matrix(X)
Y= Matrix(Y)
print(X.data)
print(Y.data)
X_t = X.transponing()
print(X_t.multiply(X).data)
print(X_t.multiply(X).reverse.data)

Ans = ((X_t.multiply(X)).reverse.data).multiply(X_t).multiply(Y)
print(Ans.data)